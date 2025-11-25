"""
Create comprehensive job application tracker with conditional formatting and data validation.
Merges existing Excel tracker with CSV analytics data.
"""
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import csv
from pathlib import Path

ROOT = Path(__file__).parent.parent
OLD_TRACKER = ROOT / 'job_hunt_tracker.xlsx'
CSV_TRACKER = ROOT / 'job_applications_tracker.csv'
NEW_TRACKER = ROOT / 'job_application_master_tracker.xlsx'

# Comprehensive column schema
COLUMN_SCHEMA = [
    # Basic Information (A-H)
    {'header': 'Application Date', 'width': 12, 'type': 'date'},
    {'header': 'Company', 'width': 20, 'type': 'text'},
    {'header': 'Position', 'width': 30, 'type': 'text'},
    {'header': 'Location', 'width': 20, 'type': 'text'},
    {'header': 'Job Type', 'width': 12, 'type': 'dropdown', 'options': ['Full-Time', 'Part-Time', 'Contract', 'Temporary']},
    {'header': 'Work Mode', 'width': 12, 'type': 'dropdown', 'options': ['On-Site', 'Remote', 'Hybrid']},
    {'header': 'Posted Date', 'width': 12, 'type': 'date'},
    {'header': 'Application URL', 'width': 15, 'type': 'url'},
    
    # Compensation & Benefits (I-N)
    {'header': 'Salary Min', 'width': 12, 'type': 'currency'},
    {'header': 'Salary Max', 'width': 12, 'type': 'currency'},
    {'header': 'Salary Target', 'width': 12, 'type': 'currency'},
    {'header': 'Travel %', 'width': 10, 'type': 'percent'},
    {'header': 'Relocation', 'width': 12, 'type': 'dropdown', 'options': ['Yes', 'No', 'Unknown', 'Negotiable']},
    {'header': 'Benefits', 'width': 25, 'type': 'text'},
    
    # Application Materials (O-R)
    {'header': 'Resume Version', 'width': 30, 'type': 'text'},
    {'header': 'Cover Letter', 'width': 30, 'type': 'text'},
    {'header': 'JD File', 'width': 30, 'type': 'text'},
    {'header': 'Portfolio/Samples', 'width': 20, 'type': 'text'},
    
    # Match Analysis (S-AB)
    {'header': 'Overall Match %', 'width': 12, 'type': 'percent'},
    {'header': 'Must-Have %', 'width': 12, 'type': 'percent'},
    {'header': 'Tech Skills %', 'width': 12, 'type': 'percent'},
    {'header': 'Process %', 'width': 12, 'type': 'percent'},
    {'header': 'Leadership %', 'width': 12, 'type': 'percent'},
    {'header': 'NPI %', 'width': 12, 'type': 'percent'},
    {'header': 'Mindset %', 'width': 12, 'type': 'percent'},
    {'header': 'Logistics %', 'width': 12, 'type': 'percent'},
    {'header': 'Years Required', 'width': 12, 'type': 'number'},
    {'header': 'Years I Have', 'width': 12, 'type': 'number'},
    
    # Gaps & Advantages (AC-AD)
    {'header': 'Key Gaps', 'width': 40, 'type': 'text'},
    {'header': 'My Advantages', 'width': 40, 'type': 'text'},
    
    # Status & Tracking (AE-AL)
    {'header': 'Priority', 'width': 12, 'type': 'dropdown', 'options': ['High', 'Medium', 'Low']},
    {'header': 'Status', 'width': 15, 'type': 'dropdown', 'options': ['Not Applied', 'Applied', 'Phone Screen', 'Interview 1', 'Interview 2', 'Interview 3+', 'Offer', 'Accepted', 'Rejected', 'Withdrawn']},
    {'header': 'Rejection Reason', 'width': 25, 'type': 'text'},
    {'header': 'Days Since Applied', 'width': 12, 'type': 'formula'},
    {'header': 'Follow-Up Date', 'width': 12, 'type': 'date'},
    {'header': 'Next Action', 'width': 25, 'type': 'text'},
    {'header': 'Action Due Date', 'width': 12, 'type': 'date'},
    {'header': 'Days Until Due', 'width': 12, 'type': 'formula'},
    
    # Contacts & Communication (AM-AQ)
    {'header': 'Recruiter Name', 'width': 20, 'type': 'text'},
    {'header': 'Recruiter Email', 'width': 25, 'type': 'text'},
    {'header': 'Recruiter Phone', 'width': 15, 'type': 'text'},
    {'header': 'Hiring Manager', 'width': 20, 'type': 'text'},
    {'header': 'Interview Dates', 'width': 30, 'type': 'text'},
    
    # Outcomes (AR-AU)
    {'header': 'Offer Amount', 'width': 12, 'type': 'currency'},
    {'header': 'Offer Date', 'width': 12, 'type': 'date'},
    {'header': 'Decision Deadline', 'width': 12, 'type': 'date'},
    {'header': 'Final Decision', 'width': 15, 'type': 'dropdown', 'options': ['Pending', 'Accepted', 'Declined', 'Withdrawn']},
    
    # Notes & Learning (AV-AW)
    {'header': 'Notes', 'width': 40, 'type': 'text'},
    {'header': 'Lessons Learned', 'width': 40, 'type': 'text'},
]


def create_master_tracker():
    """Create comprehensive tracker with all features"""
    
    print(">> Creating Master Job Application Tracker...")
    
    # Load existing data
    old_wb = load_workbook(OLD_TRACKER)
    old_ws = old_wb.active
    
    csv_data = []
    if CSV_TRACKER.exists():
        with open(CSV_TRACKER, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            csv_data = list(reader)
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    
    # Freeze panes (first row and first 3 columns)
    ws.freeze_panes = 'D2'
    
    # Create headers
    print("  [*] Creating headers...")
    for idx, col_def in enumerate(COLUMN_SCHEMA, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col_def['header']
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column width
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = col_def['width']
    
    ws.row_dimensions[1].height = 40
    
    # Migrate data from old tracker
    print("  [*] Migrating existing data...")
    row_num = 2
    for old_row in range(2, old_ws.max_row + 1):
        # Map old columns to new schema
        ws.cell(row=row_num, column=1).value = old_ws.cell(row=old_row, column=4).value  # Date Applied
        if ws.cell(row=row_num, column=1).value:
            ws.cell(row=row_num, column=1).number_format = 'MM/DD/YYYY'
        ws.cell(row=row_num, column=2).value = old_ws.cell(row=old_row, column=1).value  # Company
        ws.cell(row=row_num, column=3).value = old_ws.cell(row=old_row, column=2).value  # Position
        ws.cell(row=row_num, column=4).value = old_ws.cell(row=old_row, column=14).value  # Location
        ws.cell(row=row_num, column=5).value = old_ws.cell(row=old_row, column=16).value  # Job Type
        
        # Posted Date - normalize format
        posted_date = old_ws.cell(row=old_row, column=3).value
        if posted_date:
            if isinstance(posted_date, str):
                # Try parsing string dates
                from datetime import datetime
                try:
                    posted_date = datetime.strptime(posted_date, '%m/%d/%Y')
                except ValueError:
                    try:
                        posted_date = datetime.strptime(posted_date, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        pass  # Keep original if parsing fails
            ws.cell(row=row_num, column=7).value = posted_date
            if posted_date:
                ws.cell(row=row_num, column=7).number_format = 'MM/DD/YYYY'
        
        ws.cell(row=row_num, column=8).value = old_ws.cell(row=old_row, column=5).value  # Application URL
        ws.cell(row=row_num, column=9).value = old_ws.cell(row=old_row, column=11).value  # Salary Min
        ws.cell(row=row_num, column=10).value = old_ws.cell(row=old_row, column=12).value  # Salary Max
        ws.cell(row=row_num, column=12).value = old_ws.cell(row=old_row, column=10).value  # Travel %
        ws.cell(row=row_num, column=13).value = old_ws.cell(row=old_row, column=13).value  # Relocation
        ws.cell(row=row_num, column=15).value = old_ws.cell(row=old_row, column=7).value  # Resume Version
        ws.cell(row=row_num, column=16).value = old_ws.cell(row=old_row, column=9).value  # Cover Letter
        ws.cell(row=row_num, column=17).value = old_ws.cell(row=old_row, column=8).value  # JD File
        ws.cell(row=row_num, column=32).value = old_ws.cell(row=old_row, column=6).value  # Status
        
        row_num += 1
    
    # Add CSV data (match analytics)
    print("  [*] Adding match analytics...")
    if csv_data:
        for csv_row in csv_data:
            company = csv_row.get('Company', '')
            position = csv_row.get('Position', '')
            found_row = None
            for r in range(2, ws.max_row + 1):
                ws_company = ws.cell(row=r, column=2).value
                ws_position = ws.cell(row=r, column=3).value
                if ws_company and ws_position and str(ws_company).strip() == str(company).strip() and str(ws_position).strip() == str(position).strip():
                    found_row = r
                    break
            if found_row:
                try:
                    ws.cell(row=found_row, column=19).value = float(csv_row.get('Overall Match %', 0))
                    ws.cell(row=found_row, column=20).value = float(csv_row.get('Must-Have %', 0))
                    ws.cell(row=found_row, column=21).value = float(csv_row.get('Tech %', 0))
                    ws.cell(row=found_row, column=22).value = float(csv_row.get('Process %', 0))
                    ws.cell(row=found_row, column=23).value = float(csv_row.get('Leadership %', 0))
                    ws.cell(row=found_row, column=24).value = float(csv_row.get('NPI %', 0))
                    ws.cell(row=found_row, column=25).value = float(csv_row.get('Mindset %', 0))
                    ws.cell(row=found_row, column=26).value = float(csv_row.get('Logistics %', 0))
                    ws.cell(row=found_row, column=27).value = int(float(csv_row.get('Years Req', 0)))
                    ws.cell(row=found_row, column=28).value = int(float(csv_row.get('Years Have', 0)))
                    ws.cell(row=found_row, column=29).value = csv_row.get('Key Gaps', '')
                    ws.cell(row=found_row, column=31).value = csv_row.get('Priority', '')
                except (ValueError, TypeError) as e:
                    print(f"  Warning: Error parsing CSV data for {company}: {e}")

    # Add formulas
    print("  [*] Adding smart formulas...")
    for row in range(2, ws.max_row + 1):
        # Days Since Applied (AH) = TODAY() - Application Date (A)
        ws.cell(row=row, column=34).value = f'=IF(A{row}<>"",TODAY()-A{row},"")'
        
        # Days Until Due (AL) = Action Due Date (AK) - TODAY()
        ws.cell(row=row, column=38).value = f'=IF(AK{row}<>"",AK{row}-TODAY(),"")'

    # Add data validation
    print("  [*] Setting up data validation...")
    
    # Job Type dropdown (E)
    dv_job_type = DataValidation(type="list", formula1='"Full-Time,Part-Time,Contract,Temporary"', allow_blank=True)
    ws.add_data_validation(dv_job_type)
    dv_job_type.add('E2:E1000')
    
    # Work Mode dropdown (F)
    dv_work_mode = DataValidation(type="list", formula1='"On-Site,Remote,Hybrid"', allow_blank=True)
    ws.add_data_validation(dv_work_mode)
    dv_work_mode.add('F2:F1000')
    
    # Relocation dropdown (M)
    dv_relocation = DataValidation(type="list", formula1='"Yes,No,Unknown,Negotiable"', allow_blank=True)
    ws.add_data_validation(dv_relocation)
    dv_relocation.add('M2:M1000')
    
    # Priority dropdown (AE)
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_priority)
    dv_priority.add('AE2:AE1000')
    
    # Status dropdown (AF)
    dv_status = DataValidation(type="list", formula1='"Not Applied,Applied,Phone Screen,Interview 1,Interview 2,Interview 3+,Offer,Accepted,Rejected,Withdrawn"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('AF2:AF1000')
    
    # Final Decision dropdown (AU)
    dv_decision = DataValidation(type="list", formula1='"Pending,Accepted,Declined,Withdrawn"', allow_blank=True)
    ws.add_data_validation(dv_decision)
    dv_decision.add('AU2:AU1000')
    
    # Add conditional formatting
    print("  [*] Applying conditional formatting...")
    
    # Priority colors (AE)
    ws.conditional_formatting.add('AE2:AE1000',
        CellIsRule(operator='equal', formula=['"High"'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    ws.conditional_formatting.add('AE2:AE1000',
        CellIsRule(operator='equal', formula=['"Medium"'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    ws.conditional_formatting.add('AE2:AE1000',
        CellIsRule(operator='equal', formula=['"Low"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    
    # Status colors (AF)
    ws.conditional_formatting.add('AF2:AF1000',
        CellIsRule(operator='equal', formula=['"Offer"'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), font=Font(bold=True)))
    ws.conditional_formatting.add('AF2:AF1000',
        CellIsRule(operator='equal', formula=['"Accepted"'], fill=PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'), font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add('AF2:AF1000',
        CellIsRule(operator='equal', formula=['"Rejected"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    ws.conditional_formatting.add('AF2:AF1000',
        CellIsRule(operator='containsText', formula=['"Interview"'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    
    # Match percentage heat map (S:AB - Overall to Logistics)
    ws.conditional_formatting.add('S2:AB1000',
        ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                      mid_type='num', mid_value=50, mid_color='FFEB84',
                      end_type='num', end_value=100, end_color='63BE7B'))
    
    # Days since applied - highlight if > 14 days (AH)
    ws.conditional_formatting.add('AH2:AH1000',
        CellIsRule(operator='greaterThan', formula=['14'], fill=PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')))
    
    # Days until due - highlight if < 3 days (AL)
    ws.conditional_formatting.add('AL2:AL1000',
        CellIsRule(operator='lessThan', formula=['3'], fill=PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid'), font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add('AL2:AL1000',
        CellIsRule(operator='lessThan', formula=['7'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    
    # Match percentage heat map (S:AB - Overall to Logistics)
    ws.conditional_formatting.add('S2:AB1000',
        ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                      mid_type='num', mid_value=50, mid_color='FFEB84',
                      end_type='num', end_value=100, end_color='63BE7B'))
    
    # Days since applied - highlight if > 14 days
    ws.conditional_formatting.add(f'AF2:AF1000',
        CellIsRule(operator='greaterThan', formula=['14'], fill=PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')))
    
    # Days until due - highlight if < 3 days
    ws.conditional_formatting.add(f'AL2:AL1000',
        CellIsRule(operator='lessThan', formula=['3'], fill=PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid'), font=Font(bold=True, color='FFFFFF')))
    ws.conditional_formatting.add(f'AL2:AL1000',
        CellIsRule(operator='lessThan', formula=['7'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
    
    # Create Summary Sheet
    print("  [*] Creating summary dashboard...")
    summary_ws = wb.create_sheet("Summary Dashboard", 0)
    
    summary_ws['A1'] = "JOB APPLICATION DASHBOARD"
    summary_ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    summary_ws.merge_cells('A1:D1')
    
    summary_ws['A3'] = "Applications Summary"
    summary_ws['A3'].font = Font(bold=True, size=12)
    
    summary_ws['A4'] = "Total Applications:"
    summary_ws['B4'] = f'=COUNTA(\'Job Applications\'!B:B)-1'
    
    summary_ws['A5'] = "High Priority:"
    summary_ws['B5'] = f'=COUNTIF(\'Job Applications\'!AE:AE,"High")'
    
    summary_ws['A6'] = "Medium Priority:"
    summary_ws['B6'] = f'=COUNTIF(\'Job Applications\'!AE:AE,"Medium")'
    
    summary_ws['A7'] = "Low Priority:"
    summary_ws['B7'] = f'=COUNTIF(\'Job Applications\'!AE:AE,"Low")'
    
    summary_ws['A9'] = "Status Breakdown"
    summary_ws['A9'].font = Font(bold=True, size=12)
    
    summary_ws['A10'] = "Applied:"
    summary_ws['B10'] = f'=COUNTIF(\'Job Applications\'!AF:AF,"Applied")'
    
    summary_ws['A11'] = "Phone Screen:"
    summary_ws['B11'] = f'=COUNTIF(\'Job Applications\'!AF:AF,"Phone Screen")'
    
    summary_ws['A12'] = "Interviews:"
    summary_ws['B12'] = f'=COUNTIFS(\'Job Applications\'!AF:AF,"Interview*")'
    
    summary_ws['A13'] = "Offers:"
    summary_ws['B13'] = f'=COUNTIF(\'Job Applications\'!AF:AF,"Offer")'
    
    summary_ws['A14'] = "Accepted:"
    summary_ws['B14'] = f'=COUNTIF(\'Job Applications\'!AF:AF,"Accepted")'
    
    summary_ws['A15'] = "Rejected:"
    summary_ws['B15'] = f'=COUNTIF(\'Job Applications\'!AF:AF,"Rejected")'
    
    summary_ws['A17'] = "Match Analysis"
    summary_ws['A17'].font = Font(bold=True, size=12)
    
    summary_ws['A18'] = "Avg Overall Match:"
    summary_ws['B18'] = f'=AVERAGE(\'Job Applications\'!S:S)'
    summary_ws['B18'].number_format = '0.0"%"'
    
    summary_ws['A19'] = "Avg Must-Have Match:"
    summary_ws['B19'] = f'=AVERAGE(\'Job Applications\'!T:T)'
    summary_ws['B19'].number_format = '0.0"%"'
    
    summary_ws['A21'] = "Pending Actions"
    summary_ws['A21'].font = Font(bold=True, size=12)
    
    summary_ws['A22'] = "Follow-Ups Due:"
    summary_ws['B22'] = f'=COUNTIFS(\'Job Applications\'!AI:AI,"<="&TODAY(),\'Job Applications\'!AI:AI,"<>")'
    
    summary_ws['A23'] = "Actions Due This Week:"
    summary_ws['B23'] = f'=COUNTIFS(\'Job Applications\'!AK:AK,"<="&TODAY()+7,\'Job Applications\'!AK:AK,">="&TODAY())'
    
    # Format summary sheet
    for row in range(4, 24):
        summary_ws[f'A{row}'].font = Font(size=10)
        summary_ws[f'B{row}'].font = Font(bold=True, size=10)
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15
    
    # Save
    wb.save(NEW_TRACKER)
    print(f"\n[SUCCESS] Master Tracker Created: {NEW_TRACKER}")
    print(f"   [*] {ws.max_row - 1} applications migrated")
    print(f"   [*] {len(COLUMN_SCHEMA)} tracking columns")
    print(f"   [*] Conditional formatting applied")
    print(f"   [*] Data validation dropdowns active")
    print(f"   [*] Summary dashboard included")
    

if __name__ == '__main__':
    create_master_tracker()
