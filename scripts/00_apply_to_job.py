"""
MASTER JOB APPLICATION AUTOMATION SYSTEM
End-to-end workflow: Job Description → Analysis → Resume → Cover Letter → Tracker Update

Usage:
    python 00_apply_to_job.py "path/to/job_description.txt" --company "3M" --position "Supply Chain Engineer"
    
    Or with JD text directly:
    python 00_apply_to_job.py --jd-text "Job description here..." --company "3M" --position "Supply Chain Engineer"
"""

import os
import shutil
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
import importlib.util

# Setup paths
ROOT = Path(__file__).parent.parent
SCRIPTS_DIR = Path(__file__).parent
DATA_DIR = ROOT / 'data'
OUTPUT_DIR = ROOT / 'outputs'
JD_DIR = ROOT / 'job_descriptions'
JD_DIR.mkdir(exist_ok=True)
APPLICATIONS_DIR = ROOT / 'Applications'

sys.path.insert(0, str(SCRIPTS_DIR))

# Import modules with numbered names
def import_script(script_name, module_name):
    spec = importlib.util.spec_from_file_location(module_name, SCRIPTS_DIR / script_name)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

# Load required modules
match_engine = import_script("04_match_engine.py", "match_engine")
resume_gen = import_script("15_generate_jd_resume.py", "resume_gen")
cover_letter_gen = import_script("14_generate_cover_letter.py", "cover_letter_gen")

# Document workflow integration (taxonomy)
try:
    # Try relative import if packaged
    from . import doc_workflow  # type: ignore
    DocumentWorkflow = doc_workflow.DocumentWorkflow
except Exception:
    # Fallback to runtime import by path
    wf_spec = importlib.util.spec_from_file_location("doc_wf", SCRIPTS_DIR / "02_document_workflow.py")
    doc_wf = importlib.util.module_from_spec(wf_spec)
    if wf_spec and wf_spec.loader:
        wf_spec.loader.exec_module(doc_wf)  # type: ignore[attr-defined]
        DocumentWorkflow = getattr(doc_wf, "DocumentWorkflow")
    else:
        DocumentWorkflow = None  # type: ignore[assignment]

# Excel handling
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Install with: pip install openpyxl")


class JobApplicationAutomation:
    """Complete automation for job applications"""
    
    def __init__(self):
        self.tracker_path = ROOT / 'job_application_master_tracker.xlsx'
        self.profile = self.load_profile()
        self.contact = self.load_contact()
        self.experience = self.load_experience()
        
    def load_profile(self):
        """Load candidate profile"""
        with open(DATA_DIR / 'profile_candidate.json', 'r') as f:
            return json.load(f)
    
    def load_contact(self):
        """Load contact info"""
        with open(DATA_DIR / 'profile_contact.json', 'r') as f:
            return json.load(f)
    
    def load_experience(self):
        """Load experience data"""
        with open(DATA_DIR / 'profile_experience.json', 'r') as f:
            return json.load(f)
    
    def save_job_description(self, jd_text, company, position):
        """Save JD to file for reference"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_company = company.replace(' ', '_').replace('/', '_')
        safe_position = position.replace(' ', '_').replace('/', '_')
        filename = f"JD_{safe_company}_{safe_position}_{timestamp}.txt"
        filepath = JD_DIR / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"Company: {company}\n")
            f.write(f"Position: {position}\n")
            f.write(f"Date Saved: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")
            f.write(jd_text)
        
        return str(filename)
    
    def analyze_job(self, jd_text, company, position, location):
        """Parse JD and compute match scores"""
        print("[*] Analyzing job description...")
        
        # Determine initial priority based on keywords
        priority = "Medium"  # Default
        
        # Parse JD
        jd = match_engine.parse_job_description(
            company=company,
            role=position,
            location=location,
            priority=priority,
            jd_text=jd_text
        )
        
        # Create candidate profile
        candidate = match_engine.CandidateProfile(
            degree=self.profile.get('degree', ''),
            years_experience=self.profile.get('years_experience', 0),
            skills=set(self.profile.get('skills', [])),
            technologies=set(self.profile.get('technologies', [])),
            methodologies=set(self.profile.get('methodologies', [])),
            achievements=self.profile.get('achievements', []),
            location_preference=self.profile.get('location', 'Flexible'),
            travel_ok=self.profile.get('travel_ok', True),
            relocation_ok=self.profile.get('relocation_ok', True)
        )
        
        # Compute match
        result = match_engine.compute_match(jd, candidate)
        
        # Convert dataclass to dict for easier handling
        # Note: All scores are already percentages (0-100) from compute_match
        return {
            'overall_percent': result.overall,
            'breakdown': {
                'must_have_percent': result.must_have_score,
                'tech_percent': result.tech_score,
                'process_percent': result.process_score,
                'leadership_percent': result.leadership_score,
                'npi_percent': result.npi_score,
                'mindset_percent': result.mindset_score,
                'logistics_percent': result.logistics_score,
            },
            'gaps': result.gaps
        }

    def extract_jd_keywords(self, jd_text, top_n=10):
        """Simple keyword extractor: frequency of non-stopword tokens."""
        import re
        STOP = {
            'and','or','the','to','of','a','in','for','with','on','by','an','at','is','as','be','this','that','will','are','our','your','we'
        }
        tokens = re.findall(r'[A-Za-z][A-Za-z\-]+', jd_text)
        freq = {}
        for t in tokens:
            tl = t.lower()
            if tl in STOP or len(tl) < 3:
                continue
            freq[tl] = freq.get(tl, 0) + 1
        ranked = sorted(freq.items(), key=lambda x: x[1], reverse=True)
        return [w for w,_ in ranked[:top_n]]

    def deduplicate_achievements(self, achievements, max_items=5):
        seen = set()
        cleaned = []
        for a in achievements:
            c = a.strip().strip('*').strip('•').strip()
            if c.lower() in seen:
                continue
            seen.add(c.lower())
            cleaned.append(c)
            if len(cleaned) >= max_items:
                break
        return cleaned

    def categorize_skills(self):
        tech = []
        process = []
        soft = []
        for s in self.profile.get('skills', []):
            sl = s.lower()
            if any(k in sl for k in ['data','automation','manufacturing','technical','equipment']):
                tech.append(s)
            elif any(k in sl for k in ['process','risk','documentation','implement','safety']):
                process.append(s)
            else:
                soft.append(s)
        return {
            'Technical Skills': tech,
            'Process & Methodologies': process,
            'Core & Soft Skills': soft
        }
    
    def determine_priority(self, match_result):
        """Determine application priority based on match scores"""
        overall = match_result['overall_percent']
        must_have = match_result['breakdown'].get('must_have_percent', 0)
        
        if overall >= 70 or must_have >= 80:
            return "High"
        elif overall >= 50 or must_have >= 60:
            return "Medium"
        else:
            return "Low"
    
    def extract_competitive_advantages(self, match_result):
        """Identify candidate's competitive advantages for this role"""
        advantages = []
        
        breakdown = match_result['breakdown']
        
        if breakdown.get('leadership_percent', 0) >= 75:
            advantages.append("Strong leadership experience")
        if breakdown.get('tech_percent', 0) >= 75:
            advantages.append("Technical expertise alignment")
        if breakdown.get('npi_percent', 0) >= 75:
            advantages.append("NPI/new product experience")
        
        # Check for unique skills
        if self.profile.get('years_experience', 0) >= 10:
            advantages.append("10+ years industry experience")
        
        if 'clearance' in str(self.profile.get('skills', [])).lower():
            advantages.append("Security clearance")
        
        return "; ".join(advantages) if advantages else "N/A"
    
    def generate_resume(self, company, position, jd_text):
        """Generate JD-driven tailored resume"""
        print("[*] Generating tailored resume...")
        
        timestamp = datetime.now().strftime('%Y%m%d')
        safe_company = company.replace(' ', '_').replace('/', '_')
        safe_position = position.replace(' ', '_').replace('/', '_')
        filename = f"{safe_company}_{safe_position}_Resume_{timestamp}.docx"
        output_path = OUTPUT_DIR / filename
        
        try:
            resume_gen.build_jd_resume(company, position, jd_text, str(output_path))
            print(f"   [SUCCESS] Resume saved: {filename}")
            return filename
        except Exception as e:
            print(f"   [WARNING] Resume generation error: {e}")
            return "ERROR: See logs"
    
    def generate_cover_letter(self, company, position, jd_text):
        """Generate JD-driven cover letter"""
        print("[*] Generating tailored cover letter...")
        timestamp = datetime.now().strftime('%Y%m%d')
        safe_company = company.replace(' ', '_').replace('/', '_')
        safe_position = position.replace(' ', '_').replace('/', '_')
        filename = f"{safe_company}_{safe_position}_CoverLetter_{timestamp}.docx"
        output_path = OUTPUT_DIR / filename
        try:
            jd_keywords = self.extract_jd_keywords(jd_text, top_n=8)
            profile_merged = self.contact.copy()
            profile_merged['years_experience'] = self.profile.get('years_experience', 0)
            cover_letter_gen.build_cover_letter(company, position, jd_text, profile_merged, str(output_path), jd_keywords=jd_keywords)
            print(f"   [SUCCESS] Cover letter saved: {filename}")
            return filename
        except Exception as e:
            print(f"   [WARNING] Cover letter generation error: {e}")
            return "ERROR: See logs"
    
    def update_tracker(self, company, position, location, jd_filename, resume_filename, 
                      cover_letter_filename, match_result, priority, travel_pct=None,
                      salary_min=None, salary_max=None, relocation=None):
        """Update master tracker (49 columns) with new application"""
        
        if not EXCEL_AVAILABLE:
            print("Warning: Excel library not available. Skipping tracker update.")
            return
        
        print("[*] Updating job tracker...")
        
        try:
            wb = load_workbook(self.tracker_path)
            ws = wb.active
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Prepare data
            today = datetime.now()
            breakdown = match_result['breakdown']
            gaps = match_result.get('gaps', [])
            gaps_text = "; ".join(gaps) if gaps else ""  # gaps is already a list of strings
            advantages = self.extract_competitive_advantages(match_result)
            
            # Map to master tracker 49-column structure
            data = {
                # Basic Information (A-H)
                'A': today,  # Application Date
                'B': company,  # Company Name
                'C': position,  # Position Title
                'D': location,  # Location
                'E': 'Full-Time',  # Job Type (default)
                'F': '',  # Work Mode (user can fill)
                'G': today,  # Job Posted Date (using today as default)
                'H': '',  # Application URL (user can fill)
                
                # Compensation & Benefits (I-N)
                'I': salary_min,  # Salary Range Min
                'J': salary_max,  # Salary Range Max
                'K': '',  # Target Salary (user can fill)
                'L': travel_pct,  # Travel %
                'M': 'Yes' if relocation else 'No' if relocation is False else 'Unknown',  # Relocation
                'N': '',  # Benefits Summary (user can fill)
                
                # Application Materials (O-R)
                'O': resume_filename,  # Resume Version Used
                'P': cover_letter_filename,  # Cover Letter Used
                'Q': jd_filename,  # Job Description File
                'R': '',  # Portfolio/Samples (user can fill)
                
                # Match Analysis (S-AB)
                'S': match_result['overall_percent'],  # Overall Match %
                'T': breakdown.get('must_have_percent', 0),  # Must-Have %
                'U': breakdown.get('tech_percent', 0),  # Technical Match %
                'V': breakdown.get('process_percent', 0),  # Process Match %
                'W': breakdown.get('leadership_percent', 0),  # Leadership Match %
                'X': breakdown.get('npi_percent', 0),  # NPI Match %
                'Y': breakdown.get('mindset_percent', 0),  # Mindset Match %
                'Z': breakdown.get('logistics_percent', 0),  # Logistics Match %
                'AA': '',  # Years Required (user can fill from JD)
                'AB': '',  # Years I Have (user can fill)
                
                # Gaps & Advantages (AC-AD)
                'AC': gaps_text[:500],  # Key Gaps (truncated to 500 chars)
                'AD': advantages[:500],  # My Competitive Advantages (truncated)
                
                # Status & Tracking (AE-AL)
                'AE': priority,  # Priority (High/Medium/Low)
                'AF': 'Applied',  # Application Status
                'AG': '',  # Rejection Reason (if applicable)
                'AH': '',  # Days Since Applied (auto-calculated by Excel formula)
                'AI': '',  # Follow-Up Date (user can set)
                'AJ': '',  # Next Action (user can fill)
                'AK': '',  # Action Due Date (user can fill)
                'AL': '',  # Days Until Due (auto-calculated by Excel formula)
                
                # Contacts & Communication (AM-AQ)
                'AM': '',  # Recruiter Name (user can fill)
                'AN': '',  # Recruiter Email (user can fill)
                'AO': '',  # Recruiter Phone (user can fill)
                'AP': '',  # Hiring Manager (user can fill)
                'AQ': '',  # Interview Dates (user can fill)
                
                # Outcomes (AR-AU)
                'AR': '',  # Offer Amount (if received)
                'AS': '',  # Offer Date (if received)
                'AT': '',  # Decision Deadline (if applicable)
                'AU': 'Pending',  # Final Decision (default)
                
                # Notes & Learning (AV-AW)
                'AV': f"Auto-generated on {today.strftime('%Y-%m-%d')}",  # Notes
                'AW': '',  # Lessons Learned (user can fill later)
            }
            
            # Write data to all 49 columns
            for col, value in data.items():
                ws[f'{col}{next_row}'] = value

            # Auto-populate follow-up scheduling if empty (AJ/AK) -> 5 days ahead
            try:
                from datetime import timedelta
                follow_up_date = datetime.now() + timedelta(days=5)
                if not ws[f'AI{next_row}'].value:
                    ws[f'AI{next_row}'] = follow_up_date
                if not ws[f'AJ{next_row}'].value:
                    ws[f'AJ{next_row}'] = 'Follow-Up Email'
                if not ws[f'AK{next_row}'].value:
                    ws[f'AK{next_row}'] = follow_up_date
            except Exception:
                pass
            
            # Apply formatting to new row
            for col in data.keys():
                cell = ws[f'{col}{next_row}']
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Save
            wb.save(self.tracker_path)
            print(f"   [SUCCESS] Tracker updated (row {next_row}) - {company} - {position}")
            
        except Exception as e:
            print(f"   [WARNING] Tracker update error: {e}")
    
    def process_application(self, jd_text, company, position, location=None, 
                           travel_pct=None, salary_min=None, salary_max=None, 
                           relocation=None, application_url=None):
        """
        Complete end-to-end application processing
        
        Args:
            jd_text: Job description text
            company: Company name
            position: Position title
            location: Job location (optional)
            travel_pct: Travel percentage (optional)
            salary_min: Minimum salary (optional)
            salary_max: Maximum salary (optional)
            relocation: Relocation assistance (optional)
            application_url: URL to apply (optional)
        """
        
        print("\n" + "="*80)
        print(f">> PROCESSING APPLICATION: {company} - {position}")
        print("="*80 + "\n")
        
        # Prepare labeled run folder: outputs/<Company>/<Position>_<timestamp>
        run_stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_company = company.replace(' ', '_').replace('/', '_')
        safe_position = position.replace(' ', '_').replace('/', '_')
        run_dir = OUTPUT_DIR / safe_company / f"{safe_position}_{run_stamp}"
        run_dir.mkdir(parents=True, exist_ok=True)

        # Step 1: Save JD
        jd_filename = self.save_job_description(jd_text, company, position)
        print(f"[SUCCESS] Job description saved: {jd_filename}\n")
        
        # Step 2: Analyze match
        match_result = self.analyze_job(jd_text, company, position, location or "Not specified")
        print(f"   Overall Match: {match_result['overall_percent']:.1f}%")
        print(f"   Must-Have Match: {match_result['breakdown'].get('must_have_percent', 0):.1f}%")
        
        # Step 3: Determine priority
        priority = self.determine_priority(match_result)
        print(f"   Priority: {priority}\n")
        
        # Step 4: Generate resume
        resume_filename = self.generate_resume(company, position, jd_text)
        
        # Step 5: Generate cover letter
        cover_letter_filename = self.generate_cover_letter(company, position, jd_text)
        
        # Step 6: Update tracker
        self.update_tracker(
            company, position, location or "Not specified", 
            jd_filename, resume_filename, cover_letter_filename,
            match_result, priority, travel_pct, salary_min, salary_max, relocation
        )

        # Copy generated artifacts into labeled run folder for user review
        try:
            # Resume
            if resume_filename and not str(resume_filename).startswith('ERROR'):
                src_resume = OUTPUT_DIR / str(resume_filename)
                if src_resume.exists():
                    shutil.copy2(src_resume, run_dir / src_resume.name)
            # Cover letter
            if cover_letter_filename and not str(cover_letter_filename).startswith('ERROR'):
                src_cl = OUTPUT_DIR / str(cover_letter_filename)
                if src_cl.exists():
                    shutil.copy2(src_cl, run_dir / src_cl.name)
            # Job description (handle absolute routed path or relative)
            jd_path = Path(jd_filename)
            if not jd_path.is_absolute():
                jd_path = JD_DIR / jd_path
            if jd_path.exists():
                shutil.copy2(jd_path, run_dir / jd_path.name)
            # Save a small summary
            summary = {
                'company': company,
                'position': position,
                'location': location or 'Not specified',
                'priority': priority,
                'match_overall': match_result.get('overall_percent'),
                'generated': {
                    'resume': str(resume_filename),
                    'cover_letter': str(cover_letter_filename),
                    'job_description': str(jd_filename),
                }
            }
            with open(run_dir / 'summary.json', 'w', encoding='utf-8') as f:
                json.dump(summary, f, indent=2)
        except Exception as e:
            print(f"   [WARNING] Could not assemble labeled run folder: {e}")
        
        # Simple remaining-items list
        missing = []
        if not application_url:
            missing.append('Application URL not recorded')
        if not cover_letter_filename or cover_letter_filename.startswith('ERROR'):
            missing.append('Cover letter generation failed')

        print("\n" + "="*80)
        print("[SUCCESS] APPLICATION PROCESSING COMPLETE!")
        print("="*80)
        print(f"\nGenerated files in: {run_dir}")
        print(f"  [*] Resume: {resume_filename}")
        print(f"  [*] Cover Letter: {cover_letter_filename}")
        print(f"  [*] Job Description: {jd_filename}")
        print(f"  [*] Tracker: job_application_master_tracker.xlsx (updated)")
        print(f"\nPriority: {priority} | Match: {match_result['overall_percent']:.1f}%\n")
        if missing:
            print("Notes:")
            for item in missing:
                print(f"  - {item}")
            print()
        
        # Return result for programmatic access (e.g., GUI)
        return {
            'success': True,
            'output_folder': str(run_dir),
            'files': {
                'resume': str(resume_filename),
                'cover_letter': str(cover_letter_filename),
                'jd': str(jd_filename)
            },
            'match': match_result,
            'priority': priority
        }
    
    def apply_to_job(self, company, position, jd_text):
        """Simplified interface for GUI - just company, position, and JD text"""
        return self.process_application(
            jd_text=jd_text,
            company=company,
            position=position
        )


def main():
    parser = argparse.ArgumentParser(description='Automated Job Application System')
    parser.add_argument('jd_file', nargs='?', help='Path to job description file')
    parser.add_argument('--jd-text', help='Job description text directly')
    parser.add_argument('--company', required=True, help='Company name')
    parser.add_argument('--position', required=True, help='Position title')
    parser.add_argument('--location', help='Job location')
    parser.add_argument('--travel', type=float, help='Travel percentage')
    parser.add_argument('--salary-min', type=float, help='Minimum salary')
    parser.add_argument('--salary-max', type=float, help='Maximum salary')
    parser.add_argument('--relocation', action='store_true', help='Relocation assistance offered')
    parser.add_argument('--url', help='Application URL')
    
    args = parser.parse_args()
    
    # Get JD text
    if args.jd_text:
        jd_text = args.jd_text
    elif args.jd_file:
        with open(args.jd_file, 'r', encoding='utf-8') as f:
            jd_text = f.read()
    else:
        print("Error: Provide either jd_file or --jd-text")
        sys.exit(1)
    
    # Run automation
    automation = JobApplicationAutomation()
    automation.process_application(
        jd_text=jd_text,
        company=args.company,
        position=args.position,
        location=args.location,
        travel_pct=args.travel,
        salary_min=args.salary_min,
        salary_max=args.salary_max,
        relocation=args.relocation,
        application_url=args.url
    )


if __name__ == '__main__':
    main()
