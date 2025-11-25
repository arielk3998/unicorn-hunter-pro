import os, json, re
from dataclasses import dataclass, asdict
from typing import Set, List

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl not installed")

JOB_HUNT_DIR = os.path.join(os.path.dirname(__file__), '..', 'Job Hunting')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'enrichment.json')

KEYWORDS = {
    'lean', 'six sigma', 'dmaic', 'value stream', 'kaizen', 'npi', 'commercialization',
    'capex', 'equipment design', 'robotics', 'laser', 'molding', 'plastic', 'automation',
    'benchmarking', 'cost savings', 'complaint', 'quality', 'yield', 'throughput'
}

METRIC_PATTERN = re.compile(r"(\\$\\d+[\\d,]*|\\d+%|\\d+\\.\\d+%|\\d+\\s*(million|billion)|\\d+\\.\\d+\\s*(million|billion))", re.IGNORECASE)

@dataclass
class Enrichment:
    keywords_found: Set[str]
    metrics: List[str]
    lines_with_keywords: List[str]


def scan_workbook(path: str) -> Enrichment:
    wb = load_workbook(path, data_only=True)
    found: Set[str] = set()
    metrics: List[str] = []
    lines: List[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if not cells:
                continue
            line = ' '.join(cells)
            low = line.lower()
            kw_hit = False
            for kw in KEYWORDS:
                if kw in low:
                    found.add(kw)
                    kw_hit = True
            for m in METRIC_PATTERN.findall(line):
                metrics.append(m[0])
            if kw_hit:
                lines.append(line.strip())
    return Enrichment(found, metrics[:50], lines[:100])


def main():
    target = os.path.join(JOB_HUNT_DIR, 'jobHunt111925.xlsx')
    if not os.path.exists(target):
        print('Workbook not found:', target)
        data = Enrichment(set(), [], [])
    else:
        data = scan_workbook(target)
    payload = asdict(data)
    payload['keywords_found'] = sorted(list(payload['keywords_found']))
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2)
    print('Enrichment written to', OUTPUT_PATH)

if __name__ == '__main__':
    main()
